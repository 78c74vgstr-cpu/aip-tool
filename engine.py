import re
import json
import time
import concurrent.futures
from threading import Lock

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
from openai import OpenAI

# ── Styles ────────────────────────────────────────────────────────────────────
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BLUE_FILL   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WHITE_FONT  = Font(color="FFFFFF", bold=True)
WRAP_ALIGN  = Alignment(wrap_text=True, vertical="top")

# ── AMP ↔ MPD column name bridge ─────────────────────────────────────────────
# AMP and MPD use different names for the same columns.
AMP_TO_MPD = {
    "TASK NUMBER":            "TASK NUMBER",
    "DESCRIPTION":            "DESCRIPTION",
    "ACCESS":                 "ACCESS",
    "PREPARATION":            "PREPARATION",
    "ZONE":                   "ZONE",
    "TASK CODE":              "TASK CODE",
    "100% THRESHOLD":         "100%\nTHRESHOLD",
    "INTERVAL":               "SAMPLE\nINTERVAL",
    "100% INTERVAL":          "100%\nINTERVAL",
    "SOURCE TASK\nREFERENCE": "SOURCE TASK\nREFERENCE",
    "REFERENCE":              "REFERENCE",
}

# SOC field names that map directly to an AMP column
SOC_TO_AMP = {
    "100% INTERVAL":  "100% INTERVAL",
    "100% THRESHOLD": "100% THRESHOLD",
    "INTERVAL":       "INTERVAL",
    "THRESHOLD":      "100% THRESHOLD",
    "TASK CODE":      "TASK CODE",
    "ZONE":           "ZONE",
    "ACCESS":         "ACCESS",
    "PREPARATION":    "PREPARATION",
    "REFERENCE":      "REFERENCE",
    "DESCRIPTION":    "DESCRIPTION",
    "TASK TITLE":     "DESCRIPTION",
}

# Keywords that identify a man-hours field (belongs in planning system, not AMP)
MH_KEYWORDS = {"MH", "MEN", "MANHOUR", "MAN-HOUR"}


# ── Helpers ───────────────────────────────────────────────────────────────────

def find_header_row(ws, max_search=20):
    best_row, best_count = 1, 0
    for row in ws.iter_rows(min_row=1, max_row=max_search):
        count = sum(1 for cell in row if cell.value not in (None, ""))
        if count > best_count:
            best_count = count
            best_row   = row[0].row
    headers = {}
    for cell in ws[best_row]:
        if cell.value not in (None, ""):
            headers[str(cell.value).strip()] = cell.column
    return best_row, headers


def build_task_index(ws, header_row, task_col_idx):
    index = {}
    for col in ws.iter_cols(min_col=task_col_idx, max_col=task_col_idx,
                            min_row=header_row + 1):
        for cell in col:
            if cell.value not in (None, ""):
                index[str(cell.value).strip()] = cell.row
    return index


def read_mpd_dataframe(file_bytes):
    raw        = pd.read_excel(file_bytes, header=None, nrows=25)
    header_idx = int(raw.notna().sum(axis=1).idxmax())
    file_bytes.seek(0)
    df         = pd.read_excel(file_bytes, header=header_idx)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def normalize(val):
    return re.sub(r'\s+', ' ', str(val or "")).strip()


# ── SOC change parser (regex — no LLM needed) ────────────────────────────────

def _parse_soc_description(description: str):
    """
    Parse all 'FIELD CHANGED FROM: X, TO: Y' patterns in a SOC description.
    Returns two lists:
      - column_updates: {amp_col: new_value}   — changes we apply automatically
      - unmappable:     [{"field","old","new","is_mh"}]  — changes needing manual action
    """
    column_updates = {}
    unmappable     = []

    clean   = re.sub(r'\n+', '\n', str(description))
    pattern = r'^([\w][^\n]{0,60}?)\s+CHANGED FROM:\s*"?(.*?)"?,?\s*TO:\s*"?(.*?)"?\s*$'

    for m in re.finditer(pattern, clean, re.IGNORECASE | re.MULTILINE):
        field   = m.group(1).strip().upper()
        old_val = m.group(2).strip().strip('"\'')
        new_val = m.group(3).strip().strip('"\'')

        if not new_val:
            continue

        if field in SOC_TO_AMP:
            column_updates[SOC_TO_AMP[field]] = new_val

        elif any(kw in field for kw in MH_KEYWORDS):
            unmappable.append({
                "field": field,
                "old":   old_val,
                "new":   new_val,
                "is_mh": True,
            })

        else:
            # Applicability, Task Note, TPS Marker, Skill, In Reference(s), etc.
            unmappable.append({
                "field": field,
                "old":   old_val,
                "new":   new_val,
                "is_mh": False,
            })

    return column_updates, unmappable


def _format_soc_note(unmappable: list) -> str:
    """
    Turn unmappable change dicts into a human-readable cell value.
    Each entry shows the exact old and new value so the operator
    never needs to open the SOC document.
    """
    lines = []
    for item in unmappable:
        field = item["field"]
        old   = item["old"]
        new   = item["new"]
        if item["is_mh"]:
            lines.append(f"• {field}: {old} → {new}  (update in planning system)")
        else:
            lines.append(f"• {field}:")
            lines.append(f"    WAS: {old}")
            lines.append(f"    NOW: {new}")
    return "\n".join(lines)


# ── LLM: used ONLY for N/D/R classification ──────────────────────────────────

_BATCH_SIZE  = 20
_TIMEOUT     = 90
_MAX_RETRIES = 3
_MAX_WORKERS = 5


def _call_llm(client, system_prompt, numbered, batch_num):
    for attempt in range(1, _MAX_RETRIES + 1):
        try:
            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
                future = ex.submit(
                    client.chat.completions.create,
                    model           = "gpt-4o-mini",
                    messages        = [
                        {"role": "system", "content": system_prompt},
                        {"role": "user",   "content": f"Process:\n\n{numbered}"},
                    ],
                    response_format = {"type": "json_object"},
                    temperature     = 0,
                )
                resp = future.result(timeout=_TIMEOUT)
            parsed = json.loads(resp.choices[0].message.content)
            return parsed.get("tasks", parsed if isinstance(parsed, list) else [])
        except concurrent.futures.TimeoutError:
            wait = 2 ** attempt
            print(f"[LLM] Batch {batch_num} timeout (attempt {attempt}), retrying in {wait}s")
            time.sleep(wait)
        except Exception as exc:
            wait = 2 ** attempt
            print(f"[LLM] Batch {batch_num} error: {exc}, retrying in {wait}s")
            time.sleep(wait)
    print(f"[LLM] Batch {batch_num} permanently failed — skipping.")
    return []


def parse_soc_with_llm(soc_file_bytes, aip_columns: list, api_key: str,
                        progress_callback=None, checkpoint_state=None) -> list:
    """
    Parse the SOC file directly — no LLM needed.
    The SOC has explicit TASK NUMBER, MVT, and DESCRIPTION columns so we
    read them directly and use regex to extract all changes.

    Returns one entry per task:
      {
        "task_id":        str,
        "mvt":            "R" | "N" | "D",
        "column_updates": {amp_col: new_value},
        "unmappable":     [{"field","old","new","is_mh"}],
      }

    The api_key / checkpoint_state / progress_callback parameters are kept
    for backwards compatibility with existing app.py callers.
    """
    soc_df         = pd.read_excel(soc_file_bytes)
    soc_df.columns = [str(c).strip() for c in soc_df.columns]

    # Detect columns
    task_col = next((c for c in soc_df.columns if "task" in c.lower() and "number" in c.lower()),
                    next((c for c in soc_df.columns if "task" in c.lower()), soc_df.columns[1]))
    mvt_col  = next((c for c in soc_df.columns if c.strip().upper() == "MVT"), None)
    desc_col = next((c for c in soc_df.columns if "desc" in c.lower()), None)

    total  = len(soc_df)
    results = []

    for i, (_, row) in enumerate(soc_df.iterrows()):
        task_id = str(row.get(task_col, "")).strip()
        if not task_id or task_id.lower() == "nan":
            continue

        mvt      = str(row.get(mvt_col, "R")).strip().upper() if mvt_col else "R"
        raw_desc = str(row.get(desc_col, "")) if desc_col else ""
        if raw_desc.lower() == "nan":
            raw_desc = ""

        column_updates, unmappable = _parse_soc_description(raw_desc)

        results.append({
            "task_id":        task_id,
            "mvt":            mvt,
            "column_updates": column_updates,
            "unmappable":     unmappable,
        })

        if progress_callback and i % 50 == 0:
            progress_callback(i + 1, total)

    if progress_callback:
        progress_callback(total, total)

    return results


# ── AIP Updater ───────────────────────────────────────────────────────────────

def update_aip(aip_file_bytes, mpd_df: pd.DataFrame,
               task_changes: list, task_id_col_name: str):
    """
    Update AIP from parsed SOC changes.

    - REV column: R = revised, N = new task, blank = no change this revision
    - R tasks: changed cells updated (YELLOW). Unmappable changes (Applicability,
               Task Note, MH etc) written in full old→new format to SOC Note (RED)
    - N tasks: full row inserted from MPD (GREEN). SOC Note shows any unmappable
               changes for that task too.
    - D tasks: silently skipped
    - Tasks not in AMP: silently skipped (not applicable)

    Returns: (workbook, change_log, flagged)
    """
    wb = openpyxl.load_workbook(aip_file_bytes)
    ws = wb.active

    header_row, aip_headers = find_header_row(ws)

    if task_id_col_name not in aip_headers:
        raise ValueError(
            f"Column '{task_id_col_name}' not found in AIP.\n"
            f"Found: {list(aip_headers.keys())}"
        )

    task_col_idx   = aip_headers[task_id_col_name]
    aip_task_index = build_task_index(ws, header_row, task_col_idx)

    # MPD lookup
    mpd_task_col = next(
        (c for c in mpd_df.columns if c.strip().lower() == task_id_col_name.strip().lower()),
        next((c for c in mpd_df.columns if "task" in c.lower()), None)
    )
    if mpd_task_col is None:
        raise ValueError(f"Task ID column not found in MPD.\nMPD columns: {list(mpd_df.columns)}")

    mpd_df               = mpd_df.copy()
    mpd_df[mpd_task_col] = mpd_df[mpd_task_col].astype(str).str.strip()
    mpd_lookup           = mpd_df.set_index(mpd_task_col).to_dict(orient="index")

    # Add REV and SOC Note columns
    max_col      = ws.max_column
    rev_col      = max_col + 1
    soc_note_col = max_col + 2

    for col_idx, label in [(rev_col, "REV"), (soc_note_col, "SOC Note")]:
        h      = ws.cell(row=header_row, column=col_idx, value=label)
        h.fill = BLUE_FILL
        h.font = WHITE_FONT

    ws.column_dimensions[get_column_letter(rev_col)].width      = 6
    ws.column_dimensions[get_column_letter(soc_note_col)].width = 55

    change_log = []
    flagged    = []

    for entry in task_changes:
        task_id        = str(entry.get("task_id", "")).strip()
        mvt            = str(entry.get("mvt", "R")).strip().upper()
        column_updates = entry.get("column_updates", {})
        unmappable     = entry.get("unmappable", [])

        if not task_id:
            continue

        # ── DELETED: skip ─────────────────────────────────────────────────────
        if mvt == "D":
            continue

        # ── NEW TASK: insert row from MPD ─────────────────────────────────────
        if mvt == "N":
            if task_id in aip_task_index:
                ws.cell(row=aip_task_index[task_id], column=rev_col,
                        value="N").alignment = Alignment(horizontal="center")
                continue

            if task_id not in mpd_lookup:
                flagged.append({"Task ID": task_id, "MVT": "N",
                                "Issue": "New task not found in MPD"})
                continue

            mpd_row_data = mpd_lookup[task_id]
            new_row      = ws.max_row + 1

            for amp_col, mpd_col in AMP_TO_MPD.items():
                if amp_col not in aip_headers:
                    continue
                val = mpd_row_data.get(mpd_col, None)
                if isinstance(val, float) and pd.isna(val):
                    val = None
                cell      = ws.cell(row=new_row, column=aip_headers[amp_col], value=val)
                cell.fill = GREEN_FILL

            # Fill any untouched columns green
            for c in range(1, soc_note_col + 1):
                cell = ws.cell(row=new_row, column=c)
                rgb  = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else "00000000"
                if rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = GREEN_FILL

            rc           = ws.cell(row=new_row, column=rev_col, value="N")
            rc.fill      = GREEN_FILL
            rc.alignment = Alignment(horizontal="center")

            nc           = ws.cell(row=new_row, column=soc_note_col)
            nc.alignment = WRAP_ALIGN
            if unmappable:
                nc.value = "NEW — review applicability\n" + _format_soc_note(unmappable)
                nc.fill  = RED_FILL
            else:
                nc.value = "NEW — review applicability"
                nc.fill  = GREEN_FILL

            aip_task_index[task_id] = new_row
            change_log.append({
                "Task ID":   task_id,
                "Column":    "— NEW ROW —",
                "Old Value": "",
                "New Value": "Inserted from MPD",
            })
            continue

        # ── REVISED TASK ──────────────────────────────────────────────────────
        if task_id not in aip_task_index:
            continue  # Not in our AMP — not applicable, skip silently

        aip_row = aip_task_index[task_id]

        # Tag REV = R
        rc           = ws.cell(row=aip_row, column=rev_col, value="R")
        rc.alignment = Alignment(horizontal="center")

        # Apply column updates from MPD (authoritative source)
        if task_id in mpd_lookup:
            mpd_row_data = mpd_lookup[task_id]
            for amp_col, new_val_from_soc in column_updates.items():
                if amp_col not in aip_headers:
                    continue
                mpd_col = AMP_TO_MPD.get(amp_col)
                # Prefer MPD value; fall back to SOC-parsed value
                if mpd_col and mpd_col in mpd_row_data:
                    new_val = mpd_row_data[mpd_col]
                    if isinstance(new_val, float) and pd.isna(new_val):
                        new_val = new_val_from_soc  # fallback to SOC
                else:
                    new_val = new_val_from_soc

                cell = ws.cell(row=aip_row, column=aip_headers[amp_col])
                if normalize(cell.value) == normalize(new_val):
                    continue  # whitespace-only difference — skip

                old_str    = normalize(cell.value)
                cell.value = new_val
                cell.fill  = YELLOW_FILL

                change_log.append({
                    "Task ID":   task_id,
                    "Column":    amp_col,
                    "Old Value": old_str,
                    "New Value": normalize(new_val),
                })

        # Write SOC Note for unmappable changes
        if unmappable:
            nc           = ws.cell(row=aip_row, column=soc_note_col)
            nc.value     = _format_soc_note(unmappable)
            nc.fill      = RED_FILL
            nc.alignment = WRAP_ALIGN

    # ── Change Log sheet ──────────────────────────────────────────────────────
    if "Change Log" in wb.sheetnames:
        del wb["Change Log"]
    log_ws = wb.create_sheet("Change Log")

    for c_idx, h in enumerate(["Task ID", "Column", "Old Value", "New Value"], 1):
        cell      = log_ws.cell(row=1, column=c_idx, value=h)
        cell.fill = BLUE_FILL
        cell.font = WHITE_FONT

    for r_idx, e in enumerate(change_log, 2):
        log_ws.cell(row=r_idx, column=1, value=e["Task ID"])
        log_ws.cell(row=r_idx, column=2, value=e["Column"])
        log_ws.cell(row=r_idx, column=3, value=e["Old Value"])
        log_ws.cell(row=r_idx, column=4, value=e["New Value"])

    for col in log_ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=10) + 4
        log_ws.column_dimensions[get_column_letter(col[0].column)].width = min(width, 60)

    # ── Legend sheet ──────────────────────────────────────────────────────────
    if "Legend" in wb.sheetnames:
        del wb["Legend"]
    leg_ws = wb.create_sheet("Legend")
    leg_ws.column_dimensions["A"].width = 22
    leg_ws.column_dimensions["B"].width = 70

    leg_ws.cell(row=1, column=1, value="Colour Legend").font = Font(bold=True, size=13)
    for i, (fill, label, desc) in enumerate([
        (YELLOW_FILL, "Yellow cell",
         "Cell updated automatically from MPD"),
        (GREEN_FILL,  "Green row",
         "New task — review applicability, delete row if not applicable to your fleet"),
        (RED_FILL,    "Red SOC Note",
         "Manual action needed — exact old/new values shown for Applicability, Task Note, MH, etc."),
    ], start=3):
        leg_ws.cell(row=i, column=1, value=label).fill = fill
        leg_ws.cell(row=i, column=2, value=desc).fill  = fill

    return wb, change_log, flagged
